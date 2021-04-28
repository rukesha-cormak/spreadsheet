import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def funct(spreadsheet):
  wb = xl.load_workbook(spreadsheet)
  sheet =wb['Sheet1']
  for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    correct_price = cell.value * 4
    correct_price_cell = sheet.cell(row, 5)
    correct_price_cell.value = correct_price
  values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=5,
            max_col=5)
  chart = BarChart()
  chart.add_data(values)
  sheet.add_chart(chart, 'f2')
  wb.save(spreadsheet)

    

spreadsheet1 = 'transactions.xlsx'
action = funct(spreadsheet1)
print(action)

